# Carrega o arquivo .xlsx
from openpyxl import load_workbook
import configparser

config = configparser.ConfigParser()
config.read('./arquivos/Config.ini')


# O load carrega o arquivo e o active permite usar os dados
wb = load_workbook('./arquivos/Enviodasatividades.xlsx')
ws = wb.active

#variavel com o maximo de atividades
NotaMedia = config.getint('openpy', 'nAtividades')/2

# Mando o maximo de linhas e colunas para variáveis
max_linha = ws.max_row
max_col = ws.max_column

# Crio as duas listas que usarei
base = list()
ra = list()

# Aqui carrego as informações que vou usar na lista "ra"
# Uso a lista "base" para poder enviar para a "ra"
for i in range(2, max_linha+1):
    # Cada informação está em uma coluna, carrego todas
    cellRa = ws.cell(row=i, column = 2)
    cellNome = ws.cell(row=i, column = 3)
    cellEmail = ws.cell(row=i, column = 4)
    
    base.append(cellRa.value)
    base.append(cellNome.value)
    base.append(cellEmail.value)
    base.append(0)

    ra.append(base[:])
    
    base.clear()

# len mostra o tamanho da lista
tamanhoLista = len(ra)

pronto = list()
compara = list()

# Cria uma lista com a quantidade de atividades enviadas
# Porém ele gera itens repetidos na lista
for i in range (0, tamanhoLista):
    compara.append(ra[i])
    
    for y in range (0, tamanhoLista):
        if compara[0][0] == ra[y][0]:
            compara[0][3] = compara[0][3] + 1 
    
    pronto.append(compara[:])
    compara.clear() 


# Como o nome ja diz remove os repetidos da lista
def remove_repetidos(l):
    lista = l
    pronta = list()
    cont = 0
    tam = len(lista)
    
    for i in range (0, tam):
        for j in range (0, len(pronta)):
            if lista[i][0][0] == pronta[j][0][0]:
                cont = cont + 1
        if cont == 0:
            pronta.append(lista[i])
        cont = 0
           
    return pronta 

# Atualiza a lista
pronto = remove_repetidos(pronto) 
    
     